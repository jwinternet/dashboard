#!/usr/bin/env python3

__author__ = "Jared Winter"
__started__ = "12/1/2020"
__revision__ = "v0.0.1"

import csv
import openpyxl
import readline
import pyyaml
import sys


def txt_import(file):
	"""
	Imports a .txt file and saves its contents as a dictionary
	"""
	new_list = []
	try:
		with open(file, "r") as lines:
			for line in lines:
				stripped_line = line.strip()
				new_list.append(stripped_line)
			return new_list
	except FileNotFoundError:
		print("\n\n " + file + " - File Not Found\n\n")
		sys.exit(0)


def txt_export(file):
	"""
	Exports a dictionary into a .txt file
	"""
	device_list = []
	with open(file, "w") as lines:
		count = 1
		for dev in device_list:
			print("\n " + str(count))
		count = count + 1
		lines.write("\n\n" + dev + " - " + results)


def excel_import():
	"""
	Imports a .xlsx file and saves its contents as a dictionary
	"""
	try:
		my_excel_file = input(" .xlsx File name: ").strip()
		my_excel_sheet = input(" .xlsx Tab name: ").strip()
		workbook = openpyxl.load_workbook(my_excel_file)
		sheet = workbook.get_sheet_by_name(my_excel_sheet)
		row_count = sheet.max_row
		row_count = row_count + 1
		excel_dict = []
		for row in range(0, row_count):
			host = sheet["A" + str(row)].value
		hosts = host.rpartition("")[0]
		hostname = sheet["A" + str(row)].value
		device_type = sheet["B" + str(row)].value
		excel_dict.append({hosts: {"hostname": hostname, "device_type": device_type}})
		return excel_dict

	except FileNotFoundError:
		return " File Not Found"


def excel_export():
	"""
	Exports a dictionary into a .xlsx file
	"""
	# Asks for values for the imported spreadsheet
	print("\n Enter values for the exported excel spreadsheet...")
	h_column = input(" Hostnames Column: ").strip().upper()
	r_column = input(" Results Column: ").strip().upper()

	wb = openpyxl.Workbook()
	sheet = wb.active
	count = 0
	for dev in device_list:
		count = count + 1
	print("\n " + str(count))
	sheet[h_column + str(count)] = dev
	sheet[r_column + str(count)] = results
	wb.save(outputted_file)
	print("\n\n Finished!\n\n\n")


def csv_import(file):
	"""
	Imports a .csv file and saves its contents as a dictionary
	"""
	try:
		new_list = []
		with open(file_name, "r") as csv_input:
			read_csv = csv.reader(csv_input, delimiter=",")
		for row in read_csv:
			line = row[0]
		new_list.append(line)
		return new_list
	except FileNotFoundError:
		return file_name + " - File Not Found"


def csv_export():
	"""
	Exports a dictionary into a .csv file
	"""
	with open(outputted_file, "w") as csv_output:
		write_csv = csv.writer(csv_output, delimiter=",")
		count = 0
		for dev in device_list:
			count += 1
		row[0] = dev
		print("\n " + str(count))
		row[1] = results
		write_csv.writerow(row)
